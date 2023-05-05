import datetime
from datetime import timedelta
import json

import openpyxl as op

from recipes import count_brew_for_tanks, beer_params


def num_tank_and_brews(num, count_brew):
    '''Получение данных из файла и привидение к нужному виду'''
    with open("db.json") as f:
        a = json.load(f)
    firs_brew = a["counts_brews"]
    last_brew = firs_brew + count_brew - 1
    a["counts_brews"] += count_brew
    with open("db.json", "w") as f:
        json.dump(a, f)
    return f'{num} ({firs_brew}-{last_brew})'


def format_yeats(list):
    '''Перевод данных к нужному виду'''
    return f'{list[0]}/{list[1]}'


def start_ferments(num):
    '''Время до начала брожения'''
    hours = num * 3
    return hours


def prepear_data(data):
    '''Подготовка данных к нужному виду'''
    count_brew = count_brew_for_tanks(int(data[1]))
    numbers = num_tank_and_brews(data[1], count_brew)
    beer_name = data[2]
    date_brew = data[3].strftime('%d.%m.%y')
    yeats = format_yeats(data[-1])
    shpunt = beer_params[beer_name]['Плотность закрытия шпунта']
    cold = beer_params[beer_name]['Плотность начала охлаждения']
    start = (data[3] + timedelta(hours=start_ferments(count_brew))).strftime('%d.%m.%y')
    return [numbers, beer_name, date_brew, yeats, shpunt, cold, start]


def filling_yeats(data, text):
    '''Заполнение поля дрожжей'''
    num = data.split('/')
    return text.replace('$', num[0]) + num[1]


def prepare_data_to_protocol(data, filling_pattern):
    '''Подготовка данных для внесения в протокол'''
    blank_protocol = []
    for index in range(len(filling_pattern)):
        if '$' in filling_pattern[index]:
            blank_protocol.append(filling_yeats(data[index], filling_pattern[index]))
        else:
            blank_protocol.append(filling_pattern[index] + data[index])
    return blank_protocol


def create_blank_protocol(data):
    '''Создание протокола брожения'''

    filename = 'blanks.xlsx'
    wb = op.load_workbook(filename)

    ws = wb['Протоколы брож']

    # -- номера ячеек для заполнения
    fields_to_fill = [ws['A1'], ws['A3'], ws['F3'], ws['J3'], ws['A5'], ws['G5'], ws['B11']]

    # -- шаблон записи в протокол
    filling_pattern_protocol = [
        'Протокол брожения в ЦКТ № ',
        '   Сорт пива: ', '   Дата варки: ',
        '$-генерация из цкт №',
        '   Плотность для закрытия шпунта: ',
        '   Плотность для включения охлаждения: ', ''
        ]

    data_ok = prepear_data(data)

    # -- получаем список готовых данных
    ready_data_for_protocol = prepare_data_to_protocol(data_ok, filling_pattern_protocol)

    # -- записываем данные в протокол
    for index in range(len(ready_data_for_protocol)):
        fields_to_fill[index].value = ready_data_for_protocol[index]
    # -- заполняем даты брожения
    for day in range(1, 21):
        date = datetime.datetime.strptime(data_ok[-1], '%d.%m.%y') + timedelta(days=day)
        ws[f'B{11 + day}'] = date.strftime('%d.%m.%y')

    wb.save('blanks1.xlsx')
