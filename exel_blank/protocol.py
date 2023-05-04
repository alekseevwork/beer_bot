from datetime import datetime, timedelta
import openpyxl as op

filename = 'blanks.xlsm'

# -- надо сделать функцию определения даты начала брожения(по номеру цкт)
start_ferments = datetime.now() + timedelta(days=1)

# -- научить бота собират данные
fake_data = [
    '21 (1-8)', 'Kellers', datetime.now().strftime('%d.%m.%y'),
    '2/21', '7', '4', start_ferments.strftime('%d.%m.%y')
    ]

wb = op.load_workbook(filename)

ws = wb['Протоколы брож']

# -- номера ячеек для заполнения
fields_to_fill = [ws['A1'], ws['A3'], ws['F3'], ws['J3'], ws['A5'], ws['G5'], ws['B11']]

# -- шаблон записи в протокол
filling_pattern_protocol = [
    'Протокол брожения в ЦКТ № ',
    '   Сорт пива: ', '   Дата варки: ',
    '$-генерация из цкт №',
    '   Плотность для закрытия шпунта: ',
    '   Плотность для включения охлаждения: ', ''
    ]

# -- Протокол брожения --------------------------------------------------------------------


def filling_yeats(data, text):
    '''Заполнение поля дрожжей'''
    num = data.split('/')
    return text.replace('$', num[0]) + num[1]


def prepare_data_to_protocol(data, filling_pattern):
    '''Подготовка данных для внесения в протокол'''
    blank_protocol = []
    for index in range(len(filling_pattern)):
        if '$' in filling_pattern[index]:
            blank_protocol.append(filling_yeats(data[index], filling_pattern[index]))
        else:
            blank_protocol.append(filling_pattern[index] + data[index])
    return blank_protocol


def get_date_start_ferments(data):
    if int(data[0].split()[0]) >= 20:
        return datetime.now() + timedelta(days=1)
    else:
        return datetime.now()


def create_blank_protocol(wb, data):

    # -- получаем список готовых данных
    ready_data_for_protocol = prepare_data_to_protocol(data, filling_pattern_protocol)

    # -- записываем данные в протокол
    for index in range(len(ready_data_for_protocol)):
        fields_to_fill[index].value = ready_data_for_protocol[index]
    # -- заполняем даты брожения
    for day in range(1, 21):
        date = start_ferments + timedelta(days=day)
        ws[f'B{11 + day}'] = date.strftime('%d.%m.%y')

    # -- продумать проверку перед сохранением файла(отправка в телеграм??)
    wb.save('blanks.xlsm')
    # -- завернуть в функцию с проверкой


if __name__ == '__main__':
    get_date_start_ferments(fake_data)
